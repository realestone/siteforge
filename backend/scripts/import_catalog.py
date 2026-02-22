"""

Import BOQ catalog from the BoQ_v4.01 template Excel into PostgreSQL.

Parses 3 sheets into 4 logical sections:

  - BoQ sheet, Source of ordering = "Lyse Tele"    → section "product"  (hardware)
  - BoQ sheet, Source of ordering = "TI contractor" → section "service"  (labor/work)
  - BoM Griptel sheet                               → section "griptel"
  - BoM Solar sheet                                  → section "solar"

Only rows with a product_code in column B are imported.
Extracts 7 fields: product_code, description, comments, ordering_hints,
                    product_category, product_subcategory, vendor.
Stores row_index + sheet_name for Excel write-back on export.

"""

import sys
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, select, text
from sqlalchemy.orm import Session

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.config import settings
from app.models.base import Base
from app.models.catalog import BOQCatalogItem, CatalogSection

# Excel column indices (0-based from row tuple)
COL_B_PRODUCT_CODE = 1
COL_C_DESCRIPTION = 2
COL_E_COMMENTS = 4
COL_H_ORDERING_HINTS = 7
COL_I_PRODUCT_CATEGORY = 8
COL_J_PRODUCT_SUBCATEGORY = 9
COL_K_SOURCE_OF_ORDERING = 10
COL_L_VENDOR = 11

# Known placeholder/noise values to skip
SKIP_PRODUCT_CODES = {"MATERIAL PROVIDED BY SUBCO"}

# Header row is always row 10, data starts at row 11
HEADER_ROW = 10
DATA_START_ROW = 11


def get_sync_url() -> str:
    return settings.database_url.replace("+asyncpg", "")


def _cell_str(row: tuple, col_idx: int) -> str | None:
    """Safely extract a stripped string from a row tuple, or None if empty."""
    if col_idx >= len(row):
        return None
    val = row[col_idx].value
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _determine_boq_section(source_of_ordering: str | None) -> CatalogSection:
    """Determine section for BoQ sheet rows based on Source of ordering (col K)."""
    if source_of_ordering and source_of_ordering.lower() == "ti contractor":
        return CatalogSection.service
    return CatalogSection.product


def _import_sheet(
    session: Session,
    ws,
    sheet_name: str,
    fixed_section: CatalogSection | None = None,
) -> dict[str, int]:
    """

    Import rows from a single sheet.

    Args:
        session: DB session
        ws: openpyxl worksheet
        sheet_name: Name for sheet_name column
        fixed_section: If set, all rows get this section (for Griptel/Solar).
                       If None, section is determined per-row from col K (for BoQ).

    Returns:
        Dict with counts: {"product": N, "service": N, "skipped": N}

    """

    counts = {"product": 0, "service": 0, "griptel": 0, "solar": 0, "skipped": 0}

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=False):
        product_code = _cell_str(row, COL_B_PRODUCT_CODE)

        # Skip rows without product code (empty/placeholder tail rows)
        if not product_code:
            continue

        # Skip known noise entries
        if product_code in SKIP_PRODUCT_CODES:
            counts["skipped"] += 1
            continue

        row_index = row[0].row
        description = _cell_str(row, COL_C_DESCRIPTION) or ""
        comments = _cell_str(row, COL_E_COMMENTS)
        ordering_hints = _cell_str(row, COL_H_ORDERING_HINTS)
        product_category = _cell_str(row, COL_I_PRODUCT_CATEGORY) or ""
        product_subcategory = _cell_str(row, COL_J_PRODUCT_SUBCATEGORY)
        vendor = _cell_str(row, COL_L_VENDOR)

        # Determine section
        if fixed_section:
            section = fixed_section
        else:
            source = _cell_str(row, COL_K_SOURCE_OF_ORDERING)
            section = _determine_boq_section(source)

        # Upsert by (row_index, sheet_name)
        existing = session.execute(
            select(BOQCatalogItem).where(
                BOQCatalogItem.row_index == row_index,
                BOQCatalogItem.sheet_name == sheet_name,
            )
        ).scalar_one_or_none()

        if existing:
            existing.product_code = product_code
            existing.description = description
            existing.comments = comments
            existing.ordering_hints = ordering_hints
            existing.product_category = product_category
            existing.product_subcategory = product_subcategory
            existing.vendor = vendor
            existing.section = section
        else:
            session.add(
                BOQCatalogItem(
                    row_index=row_index,
                    sheet_name=sheet_name,
                    section=section,
                    product_code=product_code,
                    description=description,
                    comments=comments,
                    ordering_hints=ordering_hints,
                    product_category=product_category,
                    product_subcategory=product_subcategory,
                    vendor=vendor,
                )
            )

        counts[section.value] += 1

    return counts


def import_catalog(excel_path: str | None = None):
    path = Path(excel_path or settings.catalog_excel_path)
    if not path.is_absolute():
        path = Path(__file__).resolve().parent.parent / path
    if not path.exists():
        print(f"Error: Excel file not found at {path}")
        sys.exit(1)

    print(f"Loading workbook: {path.name}")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    engine = create_engine(get_sync_url(), echo=False)

    # Ensure the enum type and table exist
    Base.metadata.create_all(engine)

    total_counts = {"product": 0, "service": 0, "griptel": 0, "solar": 0, "skipped": 0}

    with Session(engine) as session:
        # 1. BoQ sheet → product + service sections
        if "BoQ" in wb.sheetnames:
            print("\nImporting BoQ sheet...")
            counts = _import_sheet(session, wb["BoQ"], "BoQ", fixed_section=None)
            for k, v in counts.items():
                total_counts[k] += v
            print(
                f"  Products: {counts['product']}, Services: {counts['service']}, Skipped: {counts['skipped']}"
            )

        # 2. BoM Griptel → griptel section
        if "BoM Griptel" in wb.sheetnames:
            print("\nImporting BoM Griptel sheet...")
            counts = _import_sheet(
                session,
                wb["BoM Griptel"],
                "BoM Griptel",
                fixed_section=CatalogSection.griptel,
            )
            for k, v in counts.items():
                total_counts[k] += v
            print(f"  Griptel items: {counts['griptel']}, Skipped: {counts['skipped']}")

        # 3. BoM Solar → solar section
        if "BoM Solar" in wb.sheetnames:
            print("\nImporting BoM Solar sheet...")
            counts = _import_sheet(
                session,
                wb["BoM Solar"],
                "BoM Solar",
                fixed_section=CatalogSection.solar,
            )
            for k, v in counts.items():
                total_counts[k] += v
            print(f"  Solar items: {counts['solar']}, Skipped: {counts['skipped']}")

        session.commit()

    wb.close()

    print(f"\n{'=' * 60}")
    print(f"Import complete:")
    print(f"  Products (hardware):  {total_counts['product']}")
    print(f"  Services (labor):     {total_counts['service']}")
    print(f"  Griptel (mounting):   {total_counts['griptel']}")
    print(f"  Solar (electrical):   {total_counts['solar']}")
    total = sum(v for k, v in total_counts.items() if k != "skipped")
    print(f"  Total catalog items:  {total}")
    print(f"  Skipped:              {total_counts['skipped']}")


def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else None
    import_catalog(excel_path)


if __name__ == "__main__":
    main()
