import io
import uuid
from pathlib import Path

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from app.config import settings
from app.database import get_db
from app.models.boq import ProjectBOQItem
from app.models.catalog import BOQCatalogItem
from app.models.project import Project
from app.schemas.boq import (
    BOQComputeResponse,
    BOQItemAdd,
    BOQItemResponse,
    BOQItemUpdate,
    RadioPlanComputeRequest,
)
from app.services.boq_service import compute_and_persist

router = APIRouter(prefix="/api/projects/{project_id}/boq", tags=["boq"])


@router.post("/compute", response_model=BOQComputeResponse)
async def compute_boq(
    project_id: uuid.UUID,
    body: RadioPlanComputeRequest,
    db: AsyncSession = Depends(get_db),
):
    """Compute BOQ from parsed radio plan data.

    Runs dependency rules, resolves product codes against the catalog,
    persists results, and returns computed items + diff.
    """
    # Verify project exists
    proj_result = await db.execute(select(Project).where(Project.id == project_id))
    if not proj_result.scalar_one_or_none():
        raise HTTPException(404, "Project not found")

    result = await compute_and_persist(project_id, body, db)
    return result


@router.get("", response_model=list[BOQItemResponse])
async def get_boq(
    project_id: uuid.UUID,
    active_only: bool = Query(False, alias="activeOnly"),
    db: AsyncSession = Depends(get_db),
):
    """Get all BOQ items for a project, with catalog data resolved."""
    stmt = (
        select(ProjectBOQItem)
        .options(joinedload(ProjectBOQItem.catalog_item))
        .where(ProjectBOQItem.project_id == project_id)
    )
    if active_only:
        stmt = stmt.where(ProjectBOQItem.quantity > 0)
    stmt = stmt.order_by(ProjectBOQItem.product_code)

    result = await db.execute(stmt)
    items = result.unique().scalars().all()
    return items


@router.post("", response_model=BOQItemResponse, status_code=201)
async def add_boq_item(
    project_id: uuid.UUID,
    body: BOQItemAdd,
    db: AsyncSession = Depends(get_db),
):
    """Add a catalog item to a project's BOQ.

    Used when the engineer searches the catalog and picks an item to add.
    """
    # Verify catalog item exists
    cat_result = await db.execute(
        select(BOQCatalogItem).where(BOQCatalogItem.id == body.catalog_item_id)
    )
    catalog_item = cat_result.scalar_one_or_none()
    if not catalog_item:
        raise HTTPException(404, "Catalog item not found")

    # Check if already in project
    existing = await db.execute(
        select(ProjectBOQItem).where(
            ProjectBOQItem.project_id == project_id,
            ProjectBOQItem.catalog_item_id == body.catalog_item_id,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(409, "Item already in project BOQ")

    item = ProjectBOQItem(
        project_id=project_id,
        catalog_item_id=body.catalog_item_id,
        product_code=catalog_item.product_code,
        quantity=body.quantity,
        is_manual_override=True,
        override_note=body.override_note,
    )
    session = db
    session.add(item)
    await session.commit()

    # Re-fetch with join to return full response
    result = await session.execute(
        select(ProjectBOQItem)
        .options(joinedload(ProjectBOQItem.catalog_item))
        .where(ProjectBOQItem.id == item.id)
    )
    return result.unique().scalar_one()


@router.patch("/{item_id}", response_model=BOQItemResponse)
async def update_boq_item(
    project_id: uuid.UUID,
    item_id: uuid.UUID,
    body: BOQItemUpdate,
    db: AsyncSession = Depends(get_db),
):
    """Update quantity on a project BOQ item."""
    result = await db.execute(
        select(ProjectBOQItem)
        .options(joinedload(ProjectBOQItem.catalog_item))
        .where(
            ProjectBOQItem.id == item_id,
            ProjectBOQItem.project_id == project_id,
        )
    )
    item = result.unique().scalar_one_or_none()
    if not item:
        raise HTTPException(404, "BOQ item not found")

    if body.quantity is not None:
        item.quantity = body.quantity
        item.is_manual_override = body.is_manual_override
        item.override_note = body.override_note
    if body.actual_quantity is not None:
        item.actual_quantity = body.actual_quantity
    if body.actual_comment is not None:
        item.actual_comment = body.actual_comment
    await db.commit()
    await db.refresh(item)
    return item


@router.delete("/{item_id}", status_code=204)
async def remove_boq_item(
    project_id: uuid.UUID,
    item_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
):
    """Remove a BOQ item from a project."""
    result = await db.execute(
        select(ProjectBOQItem).where(
            ProjectBOQItem.id == item_id,
            ProjectBOQItem.project_id == project_id,
        )
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "BOQ item not found")

    await db.delete(item)
    await db.commit()


@router.get("/export")
async def export_boq(
    project_id: uuid.UUID,
    format: str = "xlsm",
    as_built: bool = Query(False, alias="as_built"),
    db: AsyncSession = Depends(get_db),
):
    """Export project BOQ by copying the template and filling in quantities.

    1. Copies the original .xlsm template (preserving macros, formatting, formulas)
    2. Writes project metadata into rows 3-6 of each sheet
    3. Writes quantities into column D at the correct row_index for each BOQ item
    4. Returns the filled file as a download
    """
    # Resolve template path
    template_path = Path(settings.catalog_excel_path)
    if not template_path.is_absolute():
        template_path = Path(__file__).resolve().parent.parent.parent / template_path
    if not template_path.exists():
        raise HTTPException(500, f"Template not found: {template_path}")

    # Get project info
    proj_result = await db.execute(select(Project).where(Project.id == project_id))
    project = proj_result.scalar_one_or_none()
    if not project:
        raise HTTPException(404, "Project not found")

    # Get all BOQ items with catalog data
    result = await db.execute(
        select(ProjectBOQItem)
        .options(joinedload(ProjectBOQItem.catalog_item))
        .where(ProjectBOQItem.project_id == project_id)
    )
    items = result.unique().scalars().all()

    # Build lookups: (sheet_name, row_index) -> quantity / actuals
    qty_map: dict[tuple[str, int], float] = {}
    act_map: dict[tuple[str, int], tuple[float | None, str | None]] = {}
    for item in items:
        cat = item.catalog_item
        if cat and cat.sheet_name and cat.row_index:
            qty_map[(cat.sheet_name, cat.row_index)] = item.quantity
            if item.actual_quantity is not None or item.actual_comment:
                act_map[(cat.sheet_name, cat.row_index)] = (
                    item.actual_quantity,
                    item.actual_comment,
                )

    # Load template — preserve VBA for .xlsm output
    wb = openpyxl.load_workbook(str(template_path), keep_vba=(format != "xlsx"))

    # Sheets to fill metadata + quantities
    DATA_SHEETS = ("BoQ", "BoM Griptel", "BoM Solar")
    COL_D_QUANTITY = 4  # Column D = Quantity
    COL_F_ACT_QTY = 6  # Column F = ACT_Qty
    COL_G_ACT_COMMENT = 7  # Column G = ACT_Comment

    for sheet_name in wb.sheetnames:
        if sheet_name not in DATA_SHEETS:
            continue
        ws = wb[sheet_name]

        # Fill project metadata (rows 3-6, column C)
        ws.cell(row=4, column=3, value=project.site_id)  # C4 = Site ID
        ws.cell(row=5, column=3, value=project.site_name)  # C5 = Site Name

        # Fill quantities at the correct row positions
        for (sname, row_idx), qty in qty_map.items():
            if sname == sheet_name and qty > 0:
                ws.cell(row=row_idx, column=COL_D_QUANTITY, value=qty)

        # Fill actuals (columns F and G) where set
        for (sname, row_idx), (act_qty, act_comment) in act_map.items():
            if sname == sheet_name:
                if act_qty is not None:
                    ws.cell(row=row_idx, column=COL_F_ACT_QTY, value=act_qty)
                if act_comment:
                    ws.cell(row=row_idx, column=COL_G_ACT_COMMENT, value=act_comment)

    # Save to in-memory buffer
    # When format=xlsx, the workbook was loaded with keep_vba=False which
    # strips VBA macros but preserves all formatting, Tables, data validation, etc.
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    if format == "xlsx":
        ext = "xlsx"
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        ext = "xlsm"
        mime = "application/vnd.ms-excel.sheet.macroEnabled.12"

    # Increment as-built version if applicable
    if as_built:
        project.as_built_boq_version += 1
        await db.commit()
        ver = project.as_built_boq_version
        filename = f"{project.site_id or 'export'}_BOQ_AsBuilt_v{ver:02d}.{ext}"
    else:
        filename = f"BoQ_{project.site_id or 'export'}.{ext}"
    return StreamingResponse(
        buf,
        media_type=mime,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )
